# Importing needed pachages
from datetime import datetime
import openpyxl
import numpy as np
from pathlib import Path

# Importing data from the file
path = str(Path.cwd())
try:
    file_directory = ("%s/src/data.xlsx" % path) # open file in linux
except:
    file_directory = ("%s\src\data.xlsx" % path) # open file in windows
workbook = openpyxl.load_workbook(file_directory)
datas = workbook.active


# Getting products
products_list = []
for col in datas.iter_cols(max_col=1, min_row=2): # Reading the first column (products names) AND Passing title of the column which is "product"
    for cell in col:        
        products_list.append(cell.value) # Adding the product's name to produtcts_list
products_list = list(set(products_list)) # Removing duplicate products from the list


# Getting Prices and Dates
for row in datas.iter_rows(min_row=2):
    product_name, product_price, product_date = row[0].value, row[1].value, (row[2].value).date() # Removing hour, minute and second from date

    # create a dictionary named as "product's name" with values of each day's price => porduct = {'date1':'price1', 'date2':'price2', ...}
    try:
        locals()[product_name][product_date] = product_price
    except:
        locals()[product_name] = {product_date:product_price}



# Choosing product by user
while True:
    print("Please choose your product: (number)")
    for i in products_list:
        print('%i. %s' % (products_list.index(i) + 1, i)) # Printing products

    # Checking if the input is valid
    try:
        choosen_product = int(input("")) - 1
        if choosen_product >= 0:
            choosen_product = products_list[choosen_product] # Converting choosen_product from number to string
            break # A valid input has been entered
        else:
            print('\nPlease enter a valid input!\n')
    except:
        print('\nPlease enter a valid input!\n')


# Choosing calculation by user
while True:
    print('\nWhat do you want about %s? (number)' % choosen_product)

    calcuations_list = ["Last price", "Last week average price", "Lask month average price", "Last week highest price", "Last month highest price", "Last week lowest price", "Last month lowest price", "Total average price", "Choose a custome date"]

    for cal in calcuations_list:
        print('%s. %s' % (calcuations_list.index(cal) + 1, cal)) # printing calculation options

    # Checking if the input is valid
    try:
        choosen_cal = int(input("")) - 1
        if choosen_cal >= 0:
            choosen_cal = calcuations_list[choosen_cal] # Converting choose_cal from number to string
            break # A valid input has been entered
        else:
            print('\nPlease enter a valid input!\n')
    except:
        print('\nPlease enter a valid input!\n')


# Putting choosen product ditails in arrays
product_dates = np.array(sorted(list(locals()[choosen_product].keys()))) # Putting product dates in a numpy array (sorted by date)

last_date = product_dates[-1]

product_tuples = []
for date in product_dates:
    product_tuples.append((date, locals()[choosen_product][date])) # (date, price)
product_prices = np.array(product_tuples) # Putting a product prices in a numpy array => (date, price)

week_pirces = product_prices[-7: ,1] # Last 7 days prices
month_prices = product_prices[-30: ,1] # Last 30 days prices

# Removing trailing zeros for printing results
def remove_trailing_zeros(number):
    return str(float(number)).strip('0').strip('.')

# Calculating Last price
if choosen_cal == "Last price":
    last_price = product_prices[np.where(product_prices == last_date)[0], 1][0] # first index of product_prices gives the row index of the price
    print("\nLast price of %s is %s$." % (choosen_product, remove_trailing_zeros(last_price)))



# Calculating Last week average price
elif choosen_cal == "Last week average price":
    week_av = week_pirces.mean() # Average of the last 7 days prices
    print("\nLast week average price of %s is %s$." % (choosen_product, remove_trailing_zeros(week_av)))


# Calculating Last month average price
elif choosen_cal == "Lask month average price":
    month_av = month_prices.mean() # Average of the last 30 days prices
    print("\nLast month average price of %s is %s$." % (choosen_product, remove_trailing_zeros(month_av)))


# Calculating Last week highest price
elif choosen_cal == "Last week highest price":
    week_max = np.max(week_pirces) # Highest price of the last 7 days
    print("\nLast week highest price of %s is %s$." % (choosen_product, remove_trailing_zeros(week_max)))


# Calculating Last month highest price
elif choosen_cal == "Last month highest price":
    month_max = np.max(month_prices) # Highest price of the last 30 days
    print("\nLast month highest price of %s is %s$." % (choosen_product, remove_trailing_zeros(month_max)))


# Calculating Last week lowest price
elif choosen_cal == "Last week lowest price":
    week_min = np.min(week_pirces) # Lowest price of the last 7 days
    print("\nLast week lowest price of %s is %s$." % (choosen_product, remove_trailing_zeros(week_min)))


# Calculating Last month lowest price
elif choosen_cal == "Last month lowest price":
    month_min = np.min(month_prices) # Lowest price of the last 30 days
    print("\nLast month lowest price of %s is %s$." % (choosen_product, remove_trailing_zeros(month_min)))


# Calculating Total average price
elif choosen_cal == "Total average price":
    total_av = product_prices[:, 1].mean()
    print("\n%s total average price is %s$." % (choosen_product, remove_trailing_zeros(total_av)))


# Calculating Custome date
elif choosen_cal == "Choose a custome date":
    # Checking fi the input is valid
    while True:
        try:
            # Getting the start date from user
            start_date = input("Enter start date (e.g: 20/1/2022): ")
            start_date = start_date.split('/')
            start_date = datetime(int(start_date[2]), int(start_date[1]), int(start_date[0])).date()
            # Getting the finish date from user
            finish_date = input("Enter finish date (e.g: 20/1/2022): ")
            finish_date =finish_date.split('/')
            finish_date = datetime(int(finish_date[2]), int(finish_date[1]), int(finish_date[0])).date()

            # Checking if the entered date is valid
            try:
                custome_dates = product_prices[(np.where(product_prices == start_date)[0], 1)[0][0] : (np.where(product_prices == finish_date)[0], 1)[0][0] + 1] # dates and prices
                custome_prices = product_prices[(np.where(product_prices == start_date)[0], 1)[0][0] : (np.where(product_prices == finish_date)[0], 1)[0][0] + 1, 1] # prices only
                break # The input is valid
            except:
                print("\nPlease enter a date that is in 'datas' file!\n")
        except:
            print('\nPlease enter a valid input!\n')
    

    print("\nDates and Prices: ")
    print(custome_dates) # Printing custome days
    print("\nHighest Price:" , remove_trailing_zeros(np.max(custome_prices)), custome_dates[np.where(custome_prices == np.max(custome_prices))[0][0], 0]) # Printing highest price with its date
    print("Lowest Price:" , remove_trailing_zeros(np.min(custome_prices)), custome_dates[np.where(custome_prices == np.min(custome_prices))[0][0], 0]) # Printing lowest price with its date
    print("Average Price:" , remove_trailing_zeros(custome_prices.mean())) # Printing the average price